import os
import glob
import re
import json
import hashlib
import chromadb
import pandas as pd
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

SUPPORTED_EXTENSIONS = [".md", ".pdf", ".txt", ".csv", ".docx", ".xlsx", ".json"]

def get_file_hash(file_path: str) -> str:
    hash_md5 = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()

def clean_text(text: str) -> str:
    return re.sub(r'\s+', ' ', text).strip()

def recursive_split_text(text: str, chunk_size: int = 600, overlap: int = 150) -> list:
    separators = ["\n\n", "\n", ". ", " ", ""]
    
    final_chunks = []
    if len(text) <= chunk_size:
        return [text]
        
    for sep in separators:
        if sep == "":
            splits = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size-overlap)]
            return splits
            
        if sep in text:
            splits = text.split(sep)
            new_splits = []
            current_chunk = ""
            
            for split in splits:
                if len(current_chunk) + len(sep) + len(split) <= chunk_size:
                    current_chunk += (sep if current_chunk else "") + split
                else:
                    if current_chunk:
                        new_splits.append(current_chunk)
                    current_chunk = split
            if current_chunk:
                new_splits.append(current_chunk)
                
            result = []
            for ns in new_splits:
                if len(ns) > chunk_size:
                    result.extend(recursive_split_text(ns, chunk_size, overlap))
                else:
                    result.append(ns)
            return result
            
    return [text]

def extract_content(file_path: str) -> list:
    ext = os.path.splitext(file_path)[1].lower()
    results = []

    try:
        if ext == ".pdf":
            try:
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    for i, page in enumerate(pdf.pages):
                        page_text = []
                        
                        
                        tables = page.extract_tables()
                        for table in tables:
                            
                            if table:
                                
                                cleaned_table = [[str(cell) if cell else "" for cell in row] for row in table]
                                
                                header = "| " + " | ".join(cleaned_table[0]) + " |"
                                separator = "| " + " | ".join(["---"] * len(cleaned_table[0])) + " |"
                                body = "\n".join(["| " + " | ".join(row) + " |" for row in cleaned_table[1:]])
                                table_md = f"\n{header}\n{separator}\n{body}\n"
                                page_text.append(table_md)

                        
                        text = page.extract_text()
                        if text:
                            page_text.append(text)
                        
                        full_page_content = "\n\n".join(page_text)
                        if full_page_content.strip():
                            results.append({"text": full_page_content, "metadata": {"page": i + 1}})
            except ImportError:
                print("pdfplumber not found, installing fallback...")
                
                reader = PdfReader(file_path)
                for i, page in enumerate(reader.pages):
                    text = page.extract_text()
                    if text:
                        text = re.sub(r'\n{3,}', '\n\n', text)
                        results.append({"text": text, "metadata": {"page": i + 1}})
                    
        elif ext == ".docx":
            doc = DocxDocument(file_path)
            
            full_text = "\n\n".join([p.text for p in doc.paragraphs if p.text.strip()])
            results.append({"text": full_text, "metadata": {}})
            
        elif ext == ".xlsx":
            wb = load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                
                if not rows:
                    continue
                    
                
                headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
                sheet_text = []
                
                
                for row_idx, row in enumerate(rows[1:], start=2):
                  
                    if not any(row):
                        continue
                        
                    row_content = []
                    for h, cell in zip(headers, row):
                        if cell is not None and str(cell).strip():
                            row_content.append(f"{h}: {cell}")
                            
                    
                    if row_content:
                        sheet_text.append(f"[Row {row_idx}] " + " | ".join(row_content))
                
                
                results.append({"text": "\n\n".join(sheet_text), "metadata": {"sheet": sheet_name}})
            wb.close()
            
        elif ext == ".csv":
            df = pd.read_csv(file_path, dtype=str).fillna("")
            
            csv_text = []
            for idx, row in df.iterrows():
                row_content = []
                for col in df.columns:
                    val = row[col]
                    if val and val.strip():
                        row_content.append(f"{col}: {val}")
                
                if row_content:
                    csv_text.append(f"[Row {idx+1}] " + " | ".join(row_content))
            
            results.append({"text": "\n\n".join(csv_text), "metadata": {}})

        elif ext == ".json":
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            text = json.dumps(data, ensure_ascii=False, indent=2)
            results.append({"text": text, "metadata": {}})
            
        elif ext in [".md", ".txt"]:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
            results.append({"text": text, "metadata": {}})
            
    except Exception as e:
        print(f"Error extracting {file_path}: {e}")
        
    return results

def ingest(data_dir: str = "data/processed", collection_name: str = "sdu_knowledge_v3"):
    from src.core.brain import SmartBrain
    
    brain = SmartBrain(collection_name)
    if not brain.ef:
        print("Embedding Function Missing (Check API Key or Ollama Connection)")
        return

            
    collection = brain.collection
    if not collection:
        print("Failed to initialize collection.")
        return
    
    existing_docs = collection.get(include=["metadatas"])
    existing_files_map = {}
    
    if existing_docs["ids"]:
        for meta in existing_docs["metadatas"]:
            if meta and "source" in meta and "file_hash" in meta:
                existing_files_map[meta["source"]] = meta["file_hash"]

    all_files = []
    for ext in SUPPORTED_EXTENSIONS:
        all_files.extend(glob.glob(os.path.join(data_dir, f"*{ext}")))
    
    print(f"Found {len(all_files)} files.")

    new_docs = []
    new_ids = []
    new_metadatas = []
    
    processed_count = 0
    skipped_count = 0

    for file_path in all_files:
        filename = os.path.basename(file_path)
        current_hash = get_file_hash(file_path)
        
        if filename in existing_files_map:
            if existing_files_map[filename] == current_hash:
                print(f"  [SKIP] {filename} (No changes)")
                skipped_count += 1
                continue
            else:
                print(f"  [UPDATE] {filename} (Changed)")
                collection.delete(where={"source": filename})
        else:
            print(f"  [NEW] {filename}")

        contents = extract_content(file_path)
        
        file_chunks_count = 0
        for item in contents:
            raw_text = item["text"]
            base_meta = item["metadata"]
            
            chunks = recursive_split_text(raw_text)
            
            for i, chunk in enumerate(chunks):
                chunk_id = f"{filename}_{base_meta.get('page', '')}_{base_meta.get('sheet', '')}_{i}"
                chunk_id = re.sub(r'[^a-zA-Z0-9_-]', '_', chunk_id)
                
                meta = {
                    "source": filename,
                    "file_hash": current_hash,
                    "chunk_index": i,
                    **base_meta
                }
                
                new_docs.append(chunk)
                new_ids.append(chunk_id)
                new_metadatas.append(meta)
                file_chunks_count += 1
        
        processed_count += 1
        print(f"    -> Added {file_chunks_count} chunks")

    if new_docs:
        print(f"\nUpserting {len(new_docs)} new chunks...")
        batch_size = 50
        total = len(new_docs)
        
        for i in range(0, total, batch_size):
            end = min(i + batch_size, total)
            try:
                collection.upsert(
                    documents=new_docs[i:end],
                    ids=new_ids[i:end],
                    metadatas=new_metadatas[i:end]
                )
                print(f"  Batch {i}-{end} upserted.")
            except Exception as e:
                print(f"  Error upserting batch {i}-{end}: {e}")
                
        print(f"Ingestion Complete. Processed: {processed_count}, Skipped: {skipped_count}")
    else:
        print("No new data to ingest.")

if __name__ == "__main__":
    ingest()
